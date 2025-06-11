'''
Extract data from Tensorboard logs and save it to csv and xlsx files
'''

import os
from itertools import chain

import csv
from tensorboard.backend.event_processing.event_accumulator import EventAccumulator, ScalarEvent
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import openpyxl.chart as xlchart
import openpyxl.utils as xlutil


LOG_PATH = "../FoodImageCode/Results/logs/5_17_single_alpha_25"

def get_log_data(dpath: str) -> "dict[str, list[tuple[int, float]]]" :
    # Get summmaries
    summary_iterators = [
        EventAccumulator(os.path.join(dpath, dname)).Reload() \
        for dname in os.listdir(dpath) if dname.startswith("event")
    ]

    # dict for event data
    # key: scalar name (tag)
    # value: tuple containing step and scalar value
    data_dict = {}
    # To track the newest data
    data_time = {}
    # Record smallest and largest step
    step_min, step_max = float("inf"), float("-inf")

    # Iterate through all summaries
    for data in summary_iterators :
        for tag in data.scalars.Keys() :
            data_dict[tag] = []
            events: "list[ScalarEvent]" = data.Scalars(tag)

            # If there are data with the same tag, add the newest data
            wall_time = events[0].wall_time
            if data_time.get(tag) is None or wall_time > data_time[tag] :
                data_dict[tag] = [(e.step, e.value) for e in events] # Step, Value
                data_time[tag] = wall_time

                # Update step
                step_min = min(step_min, data_dict[tag][0][0])
                step_max = max(step_max, data_dict[tag][-1][0])
    
    # Make sure all data have same steps
    for key, value in data_dict.items() :
        # Pad left
        temp1 = [ (i, 0) for i in range(step_min, value[0][0]) ]
        data_dict[key] = temp1 + value
        # Pad right
        temp2 = [ (i, 0) for i in range(value[-1][0] + 1, step_max + 1) ]
        value = data_dict[key]
        data_dict[key] = value + temp2

    return data_dict

def write_csv(data_dict: dict, out_path: str) :
    with open(out_path, "w", newline="") as file :
        writer = csv.writer(file)

        # Write header row
        writer.writerow(chain(("Steps",), data_dict.keys()))

        # Write step and data
        # console.print(data_dict)
        step_min = list(data_dict.values())[0][0][0]
        for i, datas in enumerate(zip(*data_dict.values()), start=step_min) :
            writer.writerow(chain((i,), (d[1] for d in datas)))

def write_worksheet(ws: Worksheet, data_dict: dict) :
    def search_key(*search) :
        for key in data_dict.keys() :
            if all(map(lambda x: x in key, search)) :
                return key
        return None
    
    # Write header
    row, col = 1, 1
    ws.cell(row, col).value = "Steps"
    for i, key in enumerate(data_dict.keys(), start=col+1) :
        ws.cell(row, i).value = key
    
    # Write step and data
    row += 1
    for i, datas in enumerate(zip(*data_dict.values()), start=row) :
        step = datas[0][0]
        ws.cell(i, col).value = step

        # Write data
        for j, d in enumerate(datas, start=col+1) :
            ws.cell(i, j).value = d[1]
    
    # Plot chart
    tag_list = list(data_dict.keys())
    step_count = len(list(data_dict.values())[0])
    step_ref = xlchart.Reference(ws, col, row, col, row + step_count - 1)
    for j, key in enumerate(data_dict.keys(), start=col+1) :
        data_ref = xlchart.Reference(ws, j, row-1, j, row + step_count - 1)
        chart = xlchart.LineChart()
        chart.set_categories(step_ref)
        chart.add_data(data_ref, titles_from_data=True)
        ws.add_chart(chart, f"{xlutil.get_column_letter(j)}{row}")

    # Plot F1 scores and loss
    chart_col = "A"
    loss_chart = xlchart.LineChart()
    for conf in ["train", "valid"] :
        # F1 scores
        f1_chart = xlchart.LineChart()
        for metric in ["macro", "micro"] :
            key = search_key(conf, metric, "f1")
            index = tag_list.index(key) + col + 1
            data_ref = xlchart.Reference(ws, index, row-1, index, row + step_count - 1)
            f1_chart.add_data(data_ref, titles_from_data=True)
        f1_chart.set_categories(step_ref)
        f1_chart.title = f"{conf} F1"
        ws.add_chart(f1_chart, f"{chart_col}10")

        # Loss
        key = search_key(conf, "total", "loss")
        index = tag_list.index(key) + col + 1
        data_ref = xlchart.Reference(ws, index, row-1, index, row + step_count - 1)
        loss_chart.add_data(data_ref, titles_from_data=True)

        chart_col = "E"
    
    loss_chart.set_categories(step_ref)
    loss_chart.title = f"Loss"
    ws.add_chart(loss_chart, f"{chart_col}15")

data_dict = get_log_data(LOG_PATH)
write_csv(data_dict, os.path.join(LOG_PATH, "out.csv"))

wb = Workbook()
ws = wb.active
write_worksheet(ws, data_dict)
wb.save(os.path.join(LOG_PATH, "out.xlsx"))
wb.close()
